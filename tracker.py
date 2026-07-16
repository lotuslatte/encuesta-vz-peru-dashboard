#!/usr/bin/env python3
"""Reconciliación ANONIMIZADA seguimiento-de-campo (Google) × respuestas KOBO.

Lee la hoja de seguimiento de Paulo desde Drive (con una CUENTA DE SERVICIO),
clasifica el Status de campo, y la cruza contra los nombres de los respondientes
de KOBO. El matching por nombre ocurre SOLO EN MEMORIA: esta función devuelve
ÚNICAMENTE CONTEOS agregados — ningún nombre ni teléfono sale de aquí.

Diseñado para correr en GitHub Actions (stdlib + google-auth/requests/openpyxl).
Si falta el secret o la hoja no está compartida, `tracker_agg()` devuelve None y
el dashboard degrada con elegancia (omite la pestaña).

El `Status` del tracker es un campo manual poco confiable: "Contesto" = contestó
el WhatsApp / se comprometió (NO que completó); solo "Hizo la encuesta" pretende
completada. La verdad de completación es KOBO.
"""
import io
import json
import os
import re
import unicodedata
from collections import defaultdict

FILE_ID = os.environ.get(
    'TRACKER_FILE_ID', '1xEFNCLqFQ38J1PN80GGggtAF8AoPI5-u')
SHEET_RE = re.compile(r'Muestra ?\d+', re.IGNORECASE)

# Vocabulario de Status (normalizado: MAYÚS sin tildes)
S_HIZO       = {'HIZO LA ENCUESTA'}
S_CONTACTADO = {'CONTESTO'}
S_MUERTO     = {'DESCONECTADO', 'SUSPENDIDO', 'CAMBIO TITULAR',
                'NUMERO EQUIVOCADO', 'REPETIDO'}
S_RECHAZO    = {'RECHAZO'}


def _norm(s):
    s = str(s or '').upper()
    return unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode()


def _toks(s):
    s = re.sub(r'[^A-Z ]', ' ', _norm(s))
    return [t for t in s.split() if t]


def _lev(a, b):
    """Distancia de edición (Levenshtein). Espeja adist() de R."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        cur = [i]
        for j, cb in enumerate(b, 1):
            cur.append(min(prev[j] + 1, cur[j - 1] + 1,
                           prev[j - 1] + (ca != cb)))
        prev = cur
    return prev[-1]


def _match(rt, rt_sorted, pool_toks, pool_sorted,
           min_share=2, min_frac=0.60, fuzzy_maxd=2):
    """Índice del mejor match en el pool para el nombre tokenizado `rt`, o None.
    (1) contención de tokens; (2) fallback fuzzy por distancia de edición."""
    if not rt:
        return None
    shares = [len(set(rt) & set(a)) for a in pool_toks]
    if shares:
        j = max(range(len(shares)), key=lambda k: shares[k])
        if shares[j] >= min_share and shares[j] / len(rt) >= min_frac:
            return j
        dists = [_lev(rt_sorted, ps) for ps in pool_sorted]
        jj = min(range(len(dists)), key=lambda k: dists[k])
        if dists[jj] <= fuzzy_maxd and shares[jj] >= 1:
            return jj
    return None


# --- Drive (cuenta de servicio) --------------------------------------------
def fetch_tracker_xlsx(file_id=None, sa_key=None):
    """Descarga el .xlsx de la hoja desde Drive con una cuenta de servicio.
    Devuelve los bytes, o None si no hay credenciales / falla el acceso."""
    file_id = file_id or FILE_ID
    sa_key = sa_key or os.environ.get('GDRIVE_SA_KEY')
    if not sa_key:
        return None
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import AuthorizedSession
        info = json.loads(sa_key)
        creds = service_account.Credentials.from_service_account_info(
            info, scopes=['https://www.googleapis.com/auth/drive.readonly'])
        sess = AuthorizedSession(creds)
        url = f'https://www.googleapis.com/drive/v3/files/{file_id}?alt=media'
        r = sess.get(url, timeout=60)
        r.raise_for_status()
        return r.content
    except Exception as e:  # noqa: BLE001 — degradación elegante en Actions
        print(f'[tracker] no se pudo leer la hoja de Drive: {e}')
        return None


# --- Parseo del tracker -----------------------------------------------------
def parse_tracker(xlsx_bytes):
    """Devuelve filas persona-nivel de las pestañas Muestra N con Status poblado.
    Cada fila: {nombre, tanda, status, categoria, hizo, contactado, muerto,
    rechazo, enviada}. Deduplicado por nombre normalizado."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    rows = []
    for sh in wb.sheetnames:
        if not SHEET_RE.search(sh):
            continue
        ws = wb[sh]
        it = ws.iter_rows(values_only=True)
        try:
            hdr = [str(c).strip() if c is not None else '' for c in next(it)]
        except StopIteration:
            continue
        if 'Name' not in hdr or 'Status' not in hdr:
            continue
        ni, si = hdr.index('Name'), hdr.index('Status')
        gi = hdr.index('Given Name') if 'Given Name' in hdr else None
        sheet_rows = []
        has_status = False
        for r in it:
            nm = r[ni] if ni < len(r) else None
            if not nm or not str(nm).strip():
                continue
            st = r[si] if si < len(r) else None
            if st and str(st).strip():
                has_status = True
            sheet_rows.append({
                'nombre': str(nm),
                'tanda': sh,
                'genero': (str(r[gi]) if gi is not None and gi < len(r) and r[gi] else ''),
                'status': ('' if st is None else str(st)),
            })
        if has_status:
            rows.extend([x for x in sheet_rows if x['status'].strip()])

    # clasificar
    for x in rows:
        sn = _norm(x['status']).strip()
        x['hizo'] = sn in S_HIZO
        x['contactado'] = sn in S_CONTACTADO
        x['muerto'] = sn in S_MUERTO
        x['rechazo'] = sn in S_RECHAZO
        x['enviada'] = not x['muerto'] and not x['rechazo']
        if x['muerto']:
            x['categoria'] = 'numero muerto'
        elif x['rechazo']:
            x['categoria'] = 'rechazo'
        elif x['hizo']:
            x['categoria'] = 'hizo la encuesta (equipo)'
        elif x['contactado']:
            x['categoria'] = 'contesto/comprometido'
        else:
            x['categoria'] = 'enviada sin respuesta'

    # dedupe por nombre normalizado; gana la fila más "avanzada"
    best = {}
    rank = lambda x: (x['hizo'], x['contactado'])  # noqa: E731
    for x in rows:
        key = ' '.join(_toks(x['nombre']))
        if not key:
            continue
        if key not in best or rank(x) > rank(best[key]):
            best[key] = x
    return list(best.values())


# --- Reconciliación (solo conteos) -----------------------------------------
CAT_ORDER = ['hizo la encuesta (equipo)', 'contesto/comprometido',
             'enviada sin respuesta', 'numero muerto', 'rechazo']


def reconcile(tracker_rows, kobo_names):
    """Cruza tracker × nombres KOBO. Devuelve SOLO conteos agregados."""
    kobo_names = [n for n in kobo_names if n and n.strip()]
    rt = [_toks(n) for n in kobo_names]
    rt_s = [' '.join(sorted(t)) for t in rt]
    tt = [_toks(x['nombre']) for x in tracker_rows]
    tt_s = [' '.join(sorted(t)) for t in tt]

    in_kobo = [
        _match(tt[k], tt_s[k], rt, rt_s) is not None
        for k in range(len(tracker_rows))
    ]
    for x, ik in zip(tracker_rows, in_kobo):
        x['_in_kobo'] = ik

    # lado KOBO: ¿cada respuesta está en el tracker? ¿con qué categoría?
    kobo_cat = []
    for i in range(len(kobo_names)):
        j = _match(rt[i], rt_s[i], tt, tt_s)
        kobo_cat.append(tracker_rows[j]['categoria'] if j is not None else None)

    contactados = len(tracker_rows)
    enviados = sum(x['enviada'] for x in tracker_rows)
    en_kobo = sum(in_kobo)
    en_kobo_env = sum(1 for x in tracker_rows if x['enviada'] and x['_in_kobo'])

    # matriz categoría × en KOBO
    cat_n = defaultdict(int)
    cat_k = defaultdict(int)
    for x in tracker_rows:
        cat_n[x['categoria']] += 1
        if x['_in_kobo']:
            cat_k[x['categoria']] += 1
    cats = [c for c in CAT_ORDER if cat_n.get(c)]
    matrix = {
        'cats': cats,
        'n': [cat_n[c] for c in cats],
        'en_kobo': [cat_k[c] for c in cats],
    }

    # por tanda
    tandas = sorted({x['tanda'] for x in tracker_rows})
    by_tanda = {
        'tandas': tandas,
        'contactados': [sum(1 for x in tracker_rows if x['tanda'] == t) for t in tandas],
        'enviados': [sum(1 for x in tracker_rows if x['tanda'] == t and x['enviada']) for t in tandas],
        'en_kobo': [sum(1 for x in tracker_rows if x['tanda'] == t and x['_in_kobo']) for t in tandas],
    }

    disc_hizo_sin_kobo = sum(1 for x in tracker_rows if x['hizo'] and not x['_in_kobo'])
    disc_kobo_no_marcado = sum(
        1 for c in kobo_cat if c is not None and c != 'hizo la encuesta (equipo)')
    kobo_fuera = sum(1 for c in kobo_cat if c is None)

    return {
        'contactados': contactados,
        'enviados': enviados,
        'hizo': sum(x['hizo'] for x in tracker_rows),
        'contesto': sum(x['contactado'] for x in tracker_rows),
        'en_kobo': en_kobo,
        'en_kobo_env': en_kobo_env,
        'tasa': round(100 * en_kobo_env / enviados, 1) if enviados else 0,
        'matrix': matrix,
        'by_tanda': by_tanda,
        'disc_hizo_sin_kobo': disc_hizo_sin_kobo,
        'disc_kobo_no_marcado': disc_kobo_no_marcado,
        'kobo_fuera': kobo_fuera,
    }


def tracker_agg(kobo_names, file_id=None, sa_key=None):
    """Orquesta: descarga + parsea + reconcilia. None si no hay acceso."""
    xlsx = fetch_tracker_xlsx(file_id, sa_key)
    if not xlsx:
        return None
    try:
        rows = parse_tracker(xlsx)
        if not rows:
            return None
        return reconcile(rows, kobo_names)
    except Exception as e:  # noqa: BLE001
        print(f'[tracker] error procesando la hoja: {e}')
        return None
