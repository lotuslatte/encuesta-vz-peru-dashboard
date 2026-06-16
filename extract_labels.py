#!/usr/bin/env python3
"""Extrae mapas de etiquetas legibles del XLSForm v2 → labels.json.

Se corre UNA vez (localmente, requiere openpyxl). Produce labels.json con:
  - field2list: { nombre_pregunta -> nombre_lista }   (de la hoja survey)
  - lists:      { nombre_lista -> { codigo -> etiqueta } }  (de la hoja choices)

build_dashboard.py luego lee labels.json (solo stdlib) para mostrar etiquetas
en vez de códigos, sin depender del xlsx ni de openpyxl en el runner de Actions.
"""
import json
from pathlib import Path

import openpyxl

XLSX = (Path(__file__).resolve().parents[3]
        / '01_Instrument' / 'RDD_PTP-CE_encuesta_v2_KOBO.xlsx')
OUT = Path(__file__).resolve().parent / 'labels.json'

wb = openpyxl.load_workbook(XLSX, data_only=True)

# --- survey: field -> list_name ---
sv = wb['survey']
hdr = [c.value for c in sv[1]]
ci = {h: i for i, h in enumerate(hdr) if h}
field2list = {}
for row in sv.iter_rows(min_row=2, values_only=True):
    t = row[ci['type']]
    n = row[ci['name']]
    if not t or not n:
        continue
    parts = str(t).split()
    if parts[0] in ('select_one', 'select_multiple') and len(parts) >= 2:
        field2list[str(n)] = parts[1]

# --- choices: list_name -> {code: label} ---
ch = wb['choices']
chdr = [c.value for c in ch[1]]
di = {h: i for i, h in enumerate(chdr) if h}
lists = {}
for row in ch.iter_rows(min_row=2, values_only=True):
    ln = row[di['list_name']]
    code = row[di['name']]
    lbl = row[di['label']]
    if ln is None or code is None:
        continue
    lists.setdefault(str(ln), {})[str(code)] = ('' if lbl is None else str(lbl))

OUT.write_text(json.dumps({'field2list': field2list, 'lists': lists},
                          ensure_ascii=False, indent=0))
print(f'Escrito {OUT} — {len(field2list)} campos, {len(lists)} listas')
