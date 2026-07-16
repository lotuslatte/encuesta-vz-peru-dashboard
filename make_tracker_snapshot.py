#!/usr/bin/env python3
"""Genera tracker_agg.json — conteos ANONIMIZADOS de enviados↔KOBO para la
pestaña "Enviados vs Respuesta" del dashboard, SIN cuenta de servicio.

Es la vía fácil (opción snapshot): se corre a pedido, en una máquina que tenga
la hoja de seguimiento y el último export de KOBO. Escribe SOLO conteos —
ningún nombre ni teléfono queda en el JSON (ni en el repo).

Refresh:
  1) Actualizar la hoja de seguimiento local (descargarla del link de Drive) en
     TRACKER_XLSX (o pedirle a Claude que la baje: Drive ya autorizado).
  2) Tener el último export de KOBO (.xlsx) en la carpeta live_dashboard/.
  3) python3 make_tracker_snapshot.py   →  escribe tracker_agg.json
  4) commit + push del repo del dashboard.

Uso:
  python3 make_tracker_snapshot.py [ruta_hoja.xlsx] [ruta_export_kobo.xlsx]
"""
import glob
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl

import tracker

HERE = Path(__file__).resolve().parent
LIVE = HERE.parent   # 05_Deliverables/live_dashboard
NS = HERE.parents[2]  # 06_New_Survey

DEFAULT_TRACKER = NS / '07_Data_Quality' / 'inputs' / 'seguimiento_google.xlsx'

# columnas de nombre en el export "Spanish es" de KOBO
COL_NOM = 'Para empezar, ¿cuáles son sus nombres?'
COL_APE = '¿Y cuáles son sus apellidos?'


def latest_kobo_export():
    xs = sorted(glob.glob(str(LIVE / '*.xlsx')), key=lambda p: Path(p).stat().st_mtime)
    if not xs:
        sys.exit(f'No encuentro ningún export .xlsx de KOBO en {LIVE}')
    return xs[-1]


def kobo_names(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    ni = hdr.index(COL_NOM) if COL_NOM in hdr else None
    ai = hdr.index(COL_APE) if COL_APE in hdr else None
    if ni is None:
        sys.exit('El export de KOBO no tiene la columna de nombres esperada.')
    out = []
    for r in it:
        nm = (str(r[ni]) if r[ni] else '')
        if ai is not None and ai < len(r) and r[ai]:
            nm += ' ' + str(r[ai])
        nm = nm.strip()
        if nm:
            out.append(nm)
    return out


def main():
    tracker_xlsx = sys.argv[1] if len(sys.argv) > 1 else str(DEFAULT_TRACKER)
    kobo_xlsx = sys.argv[2] if len(sys.argv) > 2 else latest_kobo_export()
    if not Path(tracker_xlsx).exists():
        sys.exit(f'No encuentro la hoja de seguimiento: {tracker_xlsx}')

    rows = tracker.parse_tracker(open(tracker_xlsx, 'rb').read())
    names = kobo_names(kobo_xlsx)
    agg = tracker.reconcile(rows, names)
    agg['_asof'] = date.today().isoformat()
    agg['_kobo_n'] = len(names)

    out = HERE / 'tracker_agg.json'
    out.write_text(json.dumps(agg, ensure_ascii=False, indent=1), encoding='utf-8')
    print(f'OK — tracker_agg.json: contactados={agg["contactados"]}, '
          f'enviados={agg["enviados"]}, en_kobo={agg["en_kobo"]}, tasa={agg["tasa"]}% '
          f'(KOBO n={len(names)}, asof {agg["_asof"]})')
    # sanity: el JSON no debe contener nombres
    blob = out.read_text(encoding='utf-8')
    assert not any(x['nombre'] in blob for x in rows), 'FUGA DE PII en tracker_agg.json'
    print('   PII check OK — solo conteos.')


if __name__ == '__main__':
    main()
